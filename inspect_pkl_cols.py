import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('pkl_sheet.xlsx', data_only=True)
ws = wb['INPUT PKL']

print('='*80)
print('COLUMNS IN "INPUT PKL" SHEET:')
print('='*80)

for r in range(1, 6):
    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    non_empty = [(c, str(v)) for c, v in enumerate(row_vals, 1) if v is not None and str(v).strip() != '']
    print(f'Row {r}:')
    for col_idx, val in non_empty:
        print(f'  Col {col_idx:2d}: {val[:60]}')
    print('-'*50)
