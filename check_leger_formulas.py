import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb_data = openpyxl.load_workbook('rapor_sheet.xlsx', data_only=True)
ws_input = wb_data['INPUT SUMATIF']
ws_leger = wb_data['LEGER']

# Siswa 1: AHMAD FAUZI (Row 5 in INPUT SUMATIF, Row 7 in LEGER)
print('Siswa 1 AHMAD FAUZI:')
mapel_indices = [
    ('PAI', 7, 10, 11, 7),        # (Name, S1_col, Rata_col, Akhir_col, Leger_col)
    ('PKN', 12, 15, 16, 8),
    ('B.INDO', 17, 20, 21, 9),
    ('PJOK', 22, 25, 26, 10),
]

for name, s1_col, rata_col, akhir_col, leger_col in mapel_indices:
    s1 = ws_input.cell(5, s1_col).value
    s2 = ws_input.cell(5, s1_col+1).value
    s3 = ws_input.cell(5, s1_col+2).value
    rata = ws_input.cell(5, rata_col).value
    akhir = ws_input.cell(5, akhir_col).value
    leger_val = ws_leger.cell(7, leger_col).value
    
    # Calc options
    calc1 = (float(rata) + float(akhir)) / 2 if rata and akhir else None
    print(f'{name:8s}: S1={s1}, S2={s2}, S3={s3} => Rata={rata}, Akhir={akhir} | Leger={leger_val} (Calc (Rata+Akhir)/2 = {calc1})')
