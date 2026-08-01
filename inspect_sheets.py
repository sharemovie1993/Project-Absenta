import openpyxl
import sys

# Force UTF-8 output encoding for console print
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('rapor_sheet.xlsx', data_only=True)

target_sheets = ['INPUT SUMATIF', 'INPUT CP', 'OLAH P5', 'AREA-WALAS', 'LEGER', 'SETTING', 'MASTER']

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print('='*70)
    print(f'SHEET: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column})')
    print('='*70)
    
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(15, ws.max_column + 1))]
        non_empty = [str(v) if v is not None else '' for v in row_vals]
        if any(non_empty):
            print(f'Row {r:2d}:', non_empty[:8])
    print('\n')
