import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('rapor_sheet.xlsx', data_only=True)

sheets_to_detail = ['INPUT CP', 'OLAH P5', 'INPUT SUMATIF']

for sheet_name in sheets_to_detail:
    ws = wb[sheet_name]
    print('='*80)
    print(f'SHEET DETAILS: {sheet_name}')
    print('='*80)
    
    # Print headers (row 1-5)
    for r in range(1, min(6, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        non_empty = [(c, str(v)) for c, v in enumerate(row_vals, 1) if v is not None and str(v).strip() != '']
        if non_empty:
            print(f'Row {r}:', non_empty[:15])
            
    print('\nData Row 1 (Siswa 1):')
    for r in range(6, 7):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        non_empty = [(c, str(v)[:40]) for c, v in enumerate(row_vals, 1) if v is not None and str(v).strip() != '']
        print(f'Row {r}:', non_empty[:15])
    print('\n')
