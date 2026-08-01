import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('pkl_sheet.xlsx', data_only=True)

target_sheets = ['INPUT PKL', 'Setting Deskripsi', 'SERTIFIKAT']

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print('='*80)
    print(f'SHEET: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column})')
    print('='*80)
    
    # Print first 10 rows
    for r in range(1, min(12, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(25, ws.max_column + 1))]
        non_empty = [(c, str(v)[:40]) for c, v in enumerate(row_vals, 1) if v is not None and str(v).strip() != '']
        if non_empty:
            print(f'Row {r:2d}:', non_empty[:12])
    print('\n')
