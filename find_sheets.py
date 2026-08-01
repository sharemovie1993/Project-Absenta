import urllib.request
import re
import time

SPREADSHEET_ID = '1l4oEEPmaA5bYsnlZkxW5CcxHsb7UiNmFbZSTH1SjnDM'

# Export seluruh workbook sebagai xlsx dulu untuk mendapatkan nama sheet
xlsx_url = f'https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx'
req = urllib.request.Request(xlsx_url, headers={'User-Agent': 'Mozilla/5.0'})

print('Downloading xlsx...')
with urllib.request.urlopen(req) as response:
    xlsx_data = response.read()

with open('rapor_sheet.xlsx', 'wb') as f:
    f.write(xlsx_data)

print(f'Downloaded xlsx: {len(xlsx_data)} bytes')

# Read xlsx dengan openpyxl
try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

wb = openpyxl.load_workbook('rapor_sheet.xlsx', data_only=True)
print(f'\nTotal sheets: {len(wb.sheetnames)}')
print('Sheet names:')
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f'  [{i+1}] {name} (rows={ws.max_row}, cols={ws.max_column})')
