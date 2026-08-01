import urllib.request
import sys

SPREADSHEET_ID = '1T65Y4bx97l-NytGXbJfm2Zjiyw28lfjQCcsVE7GJIfo'
xlsx_url = f'https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx'

req = urllib.request.Request(xlsx_url, headers={'User-Agent': 'Mozilla/5.0'})
print('Downloading PKL spreadsheet xlsx...')

with urllib.request.urlopen(req) as response:
    xlsx_data = response.read()

with open('pkl_sheet.xlsx', 'wb') as f:
    f.write(xlsx_data)

print(f'Downloaded pkl_sheet.xlsx: {len(xlsx_data)} bytes')

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

wb = openpyxl.load_workbook('pkl_sheet.xlsx', data_only=True)
print(f'\nTotal sheets: {len(wb.sheetnames)}')
print('Sheet names:')
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f'  [{i+1}] {name} (rows={ws.max_row}, cols={ws.max_column})')
